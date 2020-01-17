import openpyxl
# 导入openpyxl的Workbook对象
from openpyxl import Workbook
# 导入openpyxl的风格模块
from openpyxl.styles import Font, colors, Alignment
# openpyxl专门处理字母和数字的
from openpyxl.utils import get_column_letter, column_index_from_string

'''
对于openpyxl模块我还是比较熟悉的,但是我没有创建过文件
写和查的操作比较多
'''

# 1.创建文件
wb=Workbook()
# 使用sheet,active默认是第一个sheet,猜的没查过,但是实践了才敢说
sheet=wb.active
# 插入数据,再重申一遍,程序从0开始,但是Excel不是
for col in range(1,10):
    for row in range(1,1001):
        sheet.cell(row,col).value=col+row
# 保存文件,传入参数保存的文件名
# wb.save('openpyxl文件.xlsx')

# 2.打开、写入文件(放一起讲了,),这里面有个参数还是讲一下,read_only字面意思设置是否只读,大文件建议开启,节省内存
# 打开文件:load_workbook和open模块都可以
wb= openpyxl.open('openpyxl文件.xlsx')
# wb=openpyxl.load_workbook('openpyxl文件.xlsx')

# 方法1激活sheet(就是获取要操作的sheet)
# sheet=wb.active
# 方法2,索引
# sheet=wb.index(0)
# 方法3worksheets是获取所有sheet对象,给要操作sheet的索引不就完了
# sheet=wb.worksheets[0]
# 方法4sheetnames是获取所有sheet的名字,wb['sheet名']
sheet=wb[wb.sheetnames[0]]
# 创建sheet
# sheet=wb.create_sheet('新增sheet')

# 获取有效最大行
max_row=sheet.max_row
# 获取有效最大列
max_col=sheet.max_column

# 访问单一单元格
# cell(行,列)
cell_value=sheet.cell(4,5)
# 其实就是cell(4,5),只不过是通过字母列+行访问,等下讲数字和字母互转
value=sheet['A5']

# 访问多个单元格
many_value=sheet['A4:C6']
# 字母转数字
letter=get_column_letter(5)
# 数字转字母
num=column_index_from_string(letter)

# sheet.rows是一个生成器(generator),里面是所有的行数据
rows_value=sheet.rows
# sheet.columns是一个生成器(generator),里面是所有的列数据
col_values=sheet.columns

# 设置单元格风格(只用过字体,垂直、水平居中)
# 等线24号，加粗斜体，字体颜色红色
font=Font(name='等线',size=14,italic=True,color=colors.RED,bold=True)

# 合并和拆分单元格
# 所谓合并单元格，即以合并区域的左上角的那个单元格为基准，覆盖其他单元格使之称为一个大的单元格。
# 相反，拆分单元格后将这个大单元格的值返回到原来的左上角位置。具体还是看文档
sheet.merge_cells('B1:G1')
sheet.unmerge_cells('A1:C3')

# 插入数据,举个例子
for col in range(1,10):
    for row in range(1,1001):
        # 处理Excel,大多都是根据行列获取、赋值
        # sheet.cell(行,列).value=单元格值
        sheet.cell(row,col).value=col+row
        # 设置上面写好的单元格风格
        sheet.cell(row, col).font=font
        # 设置对齐方式
        sheet.cell(row, col).alignment=Alignment(horizontal='center',vertical='center')
        # 设置行高和列宽
        sheet.row_dimensions[2].height = 40
        # C列列宽
        sheet.column_dimensions['C'].width = 30

# 保存文件,不保存啥都没了
wb.save('openpyxl文件2.xlsx')

'''
实践上的性能还没有比较过,看了前几个模块,只能说是各有千秋,没说一定要用哪个模块
'''